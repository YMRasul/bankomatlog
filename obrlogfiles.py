import openpyxl as ox
from cashwithrawal import cashWith,cashWithObr
def uzs(s):
    s1 = s.rstrip().lstrip(' ').split('UZS')
    m1 = int(s1[1])
    m2 = int(s1[2])
    return m1,m2

def kol(s):
    s1 = s.split('  ')
    m1 = int(s1[0])
    m2 = int(s1[1])
    return m1,m2

def nominal(txt,n):
    nom1,nom2 = uzs(txt[n+1][20:])
    kol1,kol2 = kol(txt[n+6][25:].strip())
    return nom1,nom2,kol1,kol2

def clearCash(txt,typ):
    i = 0
    k = i
    e = True
    while (i<len(txt) and e):
        if (not typ in txt[i]):
            i = i + 1
        else:
            k = i
            e = False
    return k

def clearCashTyp(txt):
    dt = txt[0][:8]
    n = len(txt)
    nom = 0
    while(nom < n and (not 'MACHINE NO.' in  txt[nom])):
        nom = nom + 1
    tm = '00:00'
    if nom != n:
        tm = txt[nom][19:24]

    t12 = clearCash(txt, 'TYPE 1   TYPE 2')
    t34 = clearCash(txt, 'TYPE 3   TYPE 4')
    dic = {}
    if (t12 > 0 and t34 > 0):
        n1,n2,k1,k2 = nominal(txt,t12)
        n3,n4,k3,k4 = nominal(txt,t34)
        dic['Dat'] = dt
        dic['Tim'] = tm
        dic['A'] = [n1,k1]
        dic['B'] = [n2,k2]
        dic['C'] = [n3,k3]
        dic['D'] = [n4,k4]
        return dic
    else:
        return None

def obrlog(dir,files,xls):
    wb = ox.Workbook()
    ws = wb.worksheets[0]

    lexcel = 1

    data = ["Дата","Время","A","Кол.","B","Кол.","C","Кол.","D","Кол.","   Сумма","  Выдано","A","B","C","D","N карты","Cбой","A","B","C","D"]
    for i, statN in enumerate(data):
        ws.cell(row=lexcel, column=i + 1).value = statN

    ln = len(files)
    l = 0

    while (l<ln):
        itogi = [0,0,0,0,0]
        with open( dir / files[l], 'r') as f:  # Открыть файл для чтения ==files[i]== булади
            txt = [line for line in f]
        dic = clearCashTyp(txt)
        if dic is not None:
            lexcel = lexcel + 1
            data = [dic['Dat'],dic['Tim'],
                dic["A"][0], dic["A"][1],
                dic["B"][0], dic["B"][1],
                dic["C"][0], dic["C"][1],
                dic["D"][0], dic["D"][1],
                dic["A"][0]*dic["A"][1]+
                dic["B"][0]*dic["B"][1]+
                dic["C"][0]*dic["C"][1]+
                dic["D"][0]*dic["D"][1],
                0, 0, 0, 0, 0, "", 0, 0, 0, 0, 0]
#                0, "A", "B", "C", "D", "", 0, "A", "B", "C", "D"]
            for i, statN in enumerate(data):
                ws.cell(row=lexcel, column = i + 1).value = statN

            # Дальше обработкс
            # = CASH WITHDRAWAL =

            m = cashWith(txt, '= CASH WITHDRAWAL =')
            mas = cashWithObr(txt,m)

            #[st1[:8],st1[21:29],st2[15:31],float(d1)]
            lm = len(mas)
            li = 0
            if lm > 0:
                while li < lm:
                    #print(mas[li][0],mas[li][1],mas[li][2],mas[li][3],mas[li][4])
                    lexcel = lexcel + 1
                    # 0 - признак сбоя
                    # 1 - дата
                    # 2 - время
                    # 3 - N карты
                    # 4 - Сумма
                    # 5 - A
                    # 6 - B
                    # 7 - C
                    # 8 - D
                    if (mas[li][0]==0):
                        data = [mas[li][1], mas[li][2],
                        0,0,0,0,0,0,0,0,0,
                        mas[li][4],
                        mas[li][5], mas[li][6], mas[li][7],mas[li][8],
                        mas[li][3],0, 0, 0, 0, 0]
                    else:
                        data = [mas[li][1], mas[li][2],
                        0, 0, 0, 0, 0, 0, 0, 0, 0,
                        0, 0, 0, 0, 0, mas[li][3], mas[li][4],
                        mas[li][5], mas[li][6], mas[li][7], mas[li][8]
                                ]

                    for i, statN in enumerate(data):
                        ws.cell(row=lexcel, column=i + 1).value = statN

                    itogi[0] = itogi[0] + mas[li][4]
                    itogi[1] = itogi[1] + mas[li][5]
                    itogi[2] = itogi[2] + mas[li][6]
                    itogi[3] = itogi[3] + mas[li][7]
                    itogi[4] = itogi[4] + mas[li][8]
                    li = li + 1
            #
                data = ['Итого','',
                    dic["A"][0], dic["A"][1],
                    dic["B"][0], dic["B"][1],
                    dic["C"][0], dic["C"][1],
                    dic["D"][0], dic["D"][1],
                    dic["A"][0]*dic["A"][1]+
                    dic["B"][0]*dic["B"][1]+
                    dic["C"][0]*dic["C"][1]+
                    dic["D"][0]*dic["D"][1],
                    itogi[0],
                    itogi[1], itogi[2], itogi[3], itogi[4],
                    '', '', '', '', '', '']
                for i, statN in enumerate(data):
                    ws.cell(row=lexcel, column=i + 1).value = statN
        l = l + 1
    wb.save(xls / 'some.xlsx')
    print('3-этап obrlogfiles')
    return ln
